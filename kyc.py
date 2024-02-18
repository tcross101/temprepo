from configparser import ConfigParser
import openpyxl
import http.client
import ssl
import re
import urllib.request, urllib.parse, urllib.error
import base64
import datetime
import json
from bs4 import BeautifulSoup
import requests
from requests.auth import HTTPDigestAuth

# Configuration for access to D&B
def dnb(filename='gen.ini', section='DNB'):
    # Create a parser for D&B:
    parser = ConfigParser()
    parser.read(filename)
    # Read and return headers
    db = {}
    key = ''
    if parser.has_section(section):
        items = parser.items(section)
        for item in items:
            db[item[0]] = item[1]
        str = db['api key'] + ':' + db['api secret']
        encoded_str = base64.b64encode(bytes(str, 'utf-8'))
        unencoded_str = encoded_str.decode()
        key = 'Basic ' + unencoded_str
    else:
        raise Exception('Section {0} not found in \
        the {1} file'.format(section, filename))
    return key

# Configuration for access to Google
def google(section, filename='gen.ini'):
    # Create a parser for Google
    parser = ConfigParser()
    # Read the config file
    parser.read(filename)
    db = {}
    if parser.has_section(section):
        params = parser.items(section)
        for param in params:
            db[param[0]] = param[1]
    else:
        raise Exception('Section {0} not found in \
        the {1} file'.format(section, filename))
    return db

# Obtain authorization token, which lasts for 24 hours
# [Assuming the program is constantly running, periodically pull a new
# token and update the session header.  If not, add a method that tracks time
# to pulls a new token if needed.]
def get_token(key):
    token = None
    headers = {
        'Authorization': key,
        'Content-Type': 'application/json',
    }
    data = '{ "grant_type" : "client_credentials" }'
    try:
        r = requests.post('https://plus.dnb.com/v2/token', \
            headers = headers, data = data)
        parsed = json.loads(r.text)
        if r.status_code != 200:
            print('Returned token status code is not 200.')
        else:
            token = parsed['access_token']
    except:
        print('Unable to get authorization token.')
    return token

def address_parser(address):
    location = {}
    serviceurl = 'https://maps.googleapis.com/maps/api/geocode/json?'
    # Ignore SSL certificate errors
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    parms = {}
    parms['address'] = address
    parms['key'] = google('Google')['key']
    url = serviceurl + urllib.parse.urlencode(parms)
    r = urllib.request.urlopen(url, context=ctx)
    r = r.read().decode()
    try:
        js = json.loads(r)
    except:
        return location
    try:
        location['formatted_address'] = js['results'][0]['formatted_address']
    except:
        location['formatted_address'] = ' '
    try:
        location['lat'] = js['results'][0]['geometry']['location']['lat']
    except:
        location['lat'] = ' '
    try:
        location['lng'] = js['results'][0]['geometry']['location']['lng']
    except:
        location['lng'] = ' '
    try:
        types = js['results'][0]['address_components'][:]
    except:
        types = {}
    location['street_number'] = ' '
    location['street'] = ' '
    location['city'] = ' '
    location['county'] = ' '
    location['state'] = ' '
    location['state_abbrev'] = ' '
    location['country'] = ' '
    location['country_abbrev'] = ' '
    location['postal_code'] = ' '
    for type in types:
        if type['types'][0] == 'street_number':
            location['street_number'] = type['long_name'][:]
        if type['types'][0] == 'route':
            location['street'] = type['long_name'][:]
        if type['types'][0] == 'locality':
            location['city'] = type['long_name'][:]
        if type['types'][0] == 'administrative_area_level_2':
            location['county'] = type['long_name'][:]
        if type['types'][0] == 'administrative_area_level_1':
            location['state'] = type['long_name'][:]
            location['state_abbrev'] = type['short_name']
        if type['types'][0] == 'country':
            location['country'] = type['long_name'][:]
            location['country_abbrev'] = type['short_name']
        if type['types'][0] == 'postal_code':
            location['postal_code'] = type['long_name'][:]
    return location

def company_lookup(name, location, token):
    info = {}
    serviceurl = 'https://plus.dnb.com/v1/match/cleanseMatch?'
    auth = 'Bearer ' + token
    company = name
    street = location['street_number'] + ' ' + location['street']
    headers = {
    'accept' : 'application/json;charset=utf-8',
    'authorization' : auth
    }
    params = {
    'inLanguage' : 'en-US',
    'name' : company,
    'streetAddressLine1' : street,
    'countryISOAlpha2Code' : location['country_abbrev'],
    'postalCode' : location['postal_code'],
    'addressLocality' : location['city'],
    'addressRegion' : location['state_abbrev'],
    'addressCounty' : location['county'],
    'customerReference1' : 'katsumi_company_lookup',
    'candidateMaximumQuantity' : 25,
    'isCleanseAndStandardizedInformationRequired' : True
    }
    lookup_string = serviceurl + urllib.parse.urlencode(params)
    conn = http.client.HTTPSConnection("plus.dnb.com")
    conn.request("GET", lookup_string, headers=headers)
    r = conn.getresponse()
    r = r.read().decode()
    try:
        js = json.loads(r)
    except:
        return info
    try:
        info['id'] = js['transactionDetail']['transactionID']
    except:
        info['id'] = ' '
    try:
        info['timestamp'] = js['transactionDetail']['transactionTimestamp']
    except:
        info['timestamp'] = ' '
    try:
        info['match_number'] = js['candidatesMatchedQuantity']
    except:
        info['match_number'] = ' '
    try:
        info['1st_name'] = js['matchCandidates'][0]['organization'] \
            ['primaryName']
    except:
        info['1st_name'] = ' '
    try:
        info['1st_street'] = js['matchCandidates'][0]['organization'] \
            ['primaryAddress']['streetAddress']['line1']
    except:
        info['1st_street'] = ' '
    try:
        info['1st_city'] = js['matchCandidates'][0]['organization'] \
            ['primaryAddress']['addressLocality']['name']
    except:
        info['1st_city'] = ' '
    try:
        info['1st_state'] = js['matchCandidates'][0]['organization'] \
            ['primaryAddress']['addressRegion']['abbreviatedName']
    except:
        info['1st_state'] = ' '
    try:
        info['1st_postal'] = js['matchCandidates'][0]['organization'] \
            ['primaryAddress']['postalCode']
    except:
        info['1st_postal'] = ' '
    try:
        info['1st_country'] = js['matchCandidates'][0]['organization'] \
            ['primaryAddress']['addressCountry']['isoAlpha2Code']
    except:
        info['1st_country'] = ' '
    try:
        info['1st_duns'] = js['matchCandidates'][0]['organization']['duns']
    except:
        info['1st_duns'] = ' '
    try:
        info['1st_confidence_code'] = js['matchCandidates'][0] \
        ['matchQualityInformation']['confidenceCode']
    except:
        info['1st_confidence_code'] = ' '
    try:
        info['1st_components'] = js['matchCandidates'][0] \
        ['matchQualityInformation']['matchGradeComponentsCount']
    except:
        info['1st_components'] = ' '

    try:
        items = js['matchCandidates'][0]['organization']['telephone']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][0]['organization']['telephone'] \
                [counter]['telephoneNumber']
            counter += 1
            if counter < num:
                cell += ', '
        info['1st_telephone'] = cell
    except:
        info['1st_telephone'] = ' '
    try:
        items = js['matchCandidates'][0]['organization']['tradeStyleNames']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][0]['organization'] \
                ['tradeStyleNames'][counter]['name']
            counter += 1
            if counter < num:
                cell += ', '
        info['1st_tradestyle_names'] = cell
    except:
        info['1st_tradestyle_names'] = ' '
    try:
        items = js['matchCandidates'][0]['organization']['websiteAddress']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][0]['organization'] \
                ['websiteAddress'][counter]['url']
            counter += 1
            if counter < num:
                cell += ', '
        info['1st_website'] = cell
    except:
        info['1st_website'] = ' '
    try:
        items = js['matchCandidates'][0]['organization'] \
            ['mostSeniorPrincipals']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][0]['organization'] \
                ['mostSeniorPrincipals'][counter]['fullName']
            counter += 1
            if counter < num:
                cell += ', '
        info['1st_principals'] = cell
    except:
        info['1st_principals'] = ' '
    try:
        items = js['matchCandidates'][0]['organization'] \
            ['registrationNumbers']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][0]['organization'] \
                ['registrationNumbers'][counter]['registrationNumber']
            counter += 1
            if counter < num:
                cell += ', '
        info['1st_registration_num'] = cell
    except:
        info['1st_registration_num'] = ' '
    try:
        items = js['matchCandidates'][0]['organization'] \
            ['registrationNumbers']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][0]['organization'] \
                ['registrationNumbers'][counter]['typeDescription']
            counter += 1
            if counter < num:
                cell += ', '
        info['1st_registration_type'] = cell
    except:
        info['1st_registration_type'] = ' '
    try:
        items = js['matchCandidates'][0]['organization'] \
            ['corporateLinkage']['familytreeRolesPlayed']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][0]['organization'] \
                ['corporateLinkage']['familytreeRolesPlayed'][counter] \
                ['description']
            counter += 1
            if counter < num:
                cell += ', '
        info['1st_family'] = cell
    except:
        info['1st_family'] = ' '

    try:
        info['2nd_name'] = js['matchCandidates'][1]['organization'] \
            ['primaryName']
    except:
        info['2nd_name'] = ' '
    try:
        info['1st_street'] = js['matchCandidates'][1]['organization'] \
            ['primaryAddress']['streetAddress']['line1']
    except:
        info['2nd_street'] = ' '
    try:
        info['2nd_city'] = js['matchCandidates'][1]['organization'] \
            ['primaryAddress']['addressLocality']['name']
    except:
        info['2nd_city'] = ' '
    try:
        info['2nd_state'] = js['matchCandidates'][1]['organization'] \
            ['primaryAddress']['addressRegion']['abbreviatedName']
    except:
        info['2nd_state'] = ' '
    try:
        info['2nd_postal'] = js['matchCandidates'][1]['organization'] \
            ['primaryAddress']['postalCode']
    except:
        info['2nd_postal'] = ' '
    try:
        info['2nd_country'] = js['matchCandidates'][1]['organization'] \
            ['primaryAddress']['addressCountry']['isoAlpha2Code']
    except:
        info['2nd_country'] = ' '
    try:
        info['2nd_duns'] = js['matchCandidates'][1]['organization']['duns']
    except:
        info['2nd_duns'] = ' '
    try:
        info['2nd_confidence_code'] = js['matchCandidates'][1] \
        ['matchQualityInformation']['confidenceCode']
    except:
        info['2nd_confidence_code'] = ' '
    try:
        info['2nd_components'] = js['matchCandidates'][1] \
        ['matchQualityInformation']['matchGradeComponentsCount']
    except:
        info['2nd_components'] = ' '

    try:
        items = js['matchCandidates'][1]['organization']['telephone']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][1]['organization']['telephone'] \
                [counter]['telephoneNumber']
            counter += 1
            if counter < num:
                cell += ', '
        info['2nd_telephone'] = cell
    except:
        info['2nd_telephone'] = ' '
    try:
        items = js['matchCandidates'][1]['organization']['tradeStyleNames']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][1]['organization'] \
                ['tradeStyleNames'][counter]['name']
            counter += 1
            if counter < num:
                cell += ', '
        info['2nd_tradestyle_names'] = cell
    except:
        info['2nd_tradestyle_names'] = ' '
    try:
        items = js['matchCandidates'][1]['organization'] \
            ['websiteAddress']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][1]['organization'] \
                ['websiteAddress'][counter]['url']
            counter += 1
            if counter < num:
                cell += ', '
        info['2nd_website'] = cell
    except:
        info['2nd_website'] = ' '
    try:
        items = js['matchCandidates'][1]['organization'] \
            ['mostSeniorPrincipals']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][1]['organization'] \
                ['mostSeniorPrincipals'][counter]['fullName']
            counter += 1
            if counter < num:
                cell += ', '
        info['2nd_principals'] = cell
    except:
        info['2nd_principals'] = ' '
    try:
        items = js['matchCandidates'][1]['organization'] \
            ['registrationNumbers']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][1]['organization'] \
                ['registrationNumbers'][counter]['registrationNumber']
            counter += 1
            if counter < num:
                cell += ', '
        info['2nd_registration_num'] = cell
    except:
        info['2nd_registration_num'] = ' '
    try:
        items = js['matchCandidates'][1]['organization'] \
            ['registrationNumbers']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][1]['organization'] \
                ['registrationNumbers'][counter]['typeDescription']
            counter += 1
            if counter < num:
                cell += ', '
        info['2nd_registration_type'] = cell
    except:
        info['2nd_registration_type'] = ' '
    try:
        items = js['matchCandidates'][1]['organization'] \
            ['corporateLinkage']['familytreeRolesPlayed']
        num = len(items)
        counter = 0
        cell = ''
        for item in items:
            cell += js['matchCandidates'][1]['organization'] \
                ['corporateLinkage']['familytreeRolesPlayed'][counter] \
                ['description']
            counter += 1
            if counter < num:
                cell += ', '
        info['2nd_family'] = cell
    except:
        info['2nd_family'] = ' '
    return info

# This provides more data on parents but requires an upgrade in services
def family_lookup(duns, token):
    info = {}
    info['1st_name'] = ''
    info['2nd_name'] = ''
    info['3rd_name'] = ''
    info['4th_name'] = ''
    info['1st_duns'] = ''
    info['2nd_duns'] = ''
    info['3rd_duns'] = ''
    info['4th_duns'] = ''
    info['1st_trade'] = ''
    info['2nd_trade'] = ''
    info['3rd_trade'] = ''
    info['4th_trade'] = ''
    info['1st_sic'] = ''
    info['2nd_sic'] = ''
    info['3rd_sic'] = ''
    info['4th_sic'] = ''
    info['1st_level'] = ''
    info['2nd_level'] = ''
    info['3rd_level'] = ''
    info['4th_level'] = ''
    info['1st_start'] = ''
    info['2nd_start'] = ''
    info['3rd_start'] = ''
    info['4th_start'] = ''
    info['1st_rev'] = ''
    info['2nd_rev'] = ''
    info['3rd_rev'] = ''
    info['4th_rev'] = ''
    info['1st_emp'] = ''
    info['2nd_emp'] = ''
    info['3rd_emp'] = ''
    info['4th_emp'] = ''
    info['1st_roles'] = ''
    info['2nd_roles'] = ''
    info['3rd_roles'] = ''
    info['4th_roles'] = ''
    info['1st_parent_duns'] = ''
    info['2nd_parent_duns'] = ''
    info['3rd_parent_duns'] = ''
    info['4th_parent_duns'] = ''
    try:
        serviceurl = 'https://plus.dnb.com/v1/familyTreeUpward/' + duns + '?'
        auth = 'Bearer ' + token
        headers = {'accept' : 'application/json',
            'authorization' : auth}
        params = {'customerReference' : 'katsumi_family_lookup'}
        lookup_string = serviceurl + urllib.parse.urlencode(params)
        conn = http.client.HTTPSConnection("plus.dnb.com")
        conn.request("GET", lookup_string, headers=headers)
        r = conn.getresponse()
        r = r.read().decode()
        js = json.loads(r)
    except:
        return info
    try:
        info['id'] = js['transactionDetail']['transactionID']
    except:
        info['id'] = ' '
    try:
        info['timestamp'] = js['transactionDetail']['transactionTimestamp']
    except:
        info['timestamp'] = ' '
    try:
        members = js['familyTreeMembers']
        print(members)
        for member in members:
            level = member['corporateLinkage']['hierarchyLevel']
            if level == 1:
                info['1st_level'] = 1
                try:
                    info['1st_name'] = member['primaryName']
                except:
                    pass
                try:
                    info['1st_duns'] = member['duns']
                except:
                    pass
                try:
                    items = member['tradeStyleNames']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['name']
                        priority = items[counter]['priority']
                        cell += '('+str(priority)+')'
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['1st_trade'] = cell
                except:
                    pass
                try:
                    info['1st_sic'] = member['primaryIndustryCode']['usSicV4']
                except:
                    pass
                try:
                    info['1st_start'] = member['startDate']
                except:
                    pass
                try:
                    info['1st_emp'] = member['numberOfEmployees'][0]['value']
                except:
                    pass
                try:
                    info['1st_rev'] = member['financials'][0] \
                        ['yearlyRevenue'][0]['value']
                except:
                    pass
                try:
                    info['1st_parent_duns'] = member['corporateLinkage'] \
                        ['parent']['duns']
                except:
                    pass
                try:
                    items = member['corporateLinkage']['familytreeRolesPlayed']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['description']
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['1st_roles'] = cell
                except:
                    pass

            elif level == 2:
                info['2nd_level'] = 2
                try:
                    info['2nd_name'] = member['primaryName']
                except:
                    pass
                try:
                    info['2nd_duns'] = member['duns']
                except:
                    pass
                try:
                    items = member['tradeStyleNames']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['name']
                        priority = items[counter]['priority']
                        cell += '('+str(priority)+')'
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['2nd_trade'] = cell
                except:
                    pass
                try:
                    info['2nd_sic'] = member['primaryIndustryCode']['usSicV4']
                except:
                    pass
                try:
                    info['2nd_start'] = member['startDate']
                except:
                    pass
                try:
                    info['2nd_emp'] = member['numberOfEmployees'][0]['value']
                except:
                    pass
                try:
                    info['2nd_rev'] = member['financials'][0] \
                        ['yearlyRevenue'][0]['value']
                except:
                    pass
                try:
                    info['2nd_parent_duns'] = member['corporateLinkage'] \
                        ['parent']['duns']
                except:
                    pass
                try:
                    items = member['corporateLinkage']['familytreeRolesPlayed']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['description']
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['2nd_roles'] = cell
                except:
                    pass

            elif level == 3:
                info['3rd_level'] = 3
                try:
                    info['3rd_name'] = member['primaryName']
                except:
                    pass
                try:
                    info['3rd_duns'] = member['duns']
                except:
                    pass
                try:
                    items = member['tradeStyleNames']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['name']
                        priority = items[counter]['priority']
                        cell += '('+str(priority)+')'
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['3rd_trade'] = cell
                except:
                    pass
                try:
                    info['3rd_sic'] = member['primaryIndustryCode']['usSicV4']
                except:
                    pass
                try:
                    info['3rd_start'] = member['startDate']
                except:
                    pass
                try:
                    info['3rd_emp'] = member['numberOfEmployees'][0]['value']
                except:
                    pass
                try:
                    info['3rd_rev'] = member['financials'][0] \
                        ['yearlyRevenue'][0]['value']
                except:
                    pass
                try:
                    info['3rd_parent_duns'] = member['corporateLinkage'] \
                        ['parent']['duns']
                except:
                    pass
                try:
                    items = member['corporateLinkage']['familytreeRolesPlayed']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['description']
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['3rd_roles'] = cell
                except:
                    pass

            else:
                info['4th_level'] = 4
                try:
                    info['4th_name'] = member['primaryName']
                except:
                    pass
                try:
                    info['4th_duns'] = member['duns']
                except:
                    pass
                try:
                    items = member['tradeStyleNames']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['name']
                        priority = items[counter]['priority']
                        cell += '('+str(priority)+')'
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['4th_trade'] = cell
                except:
                    pass
                try:
                    info['4th_sic'] = member['primaryIndustryCode']['usSicV4']
                except:
                    pass
                try:
                    info['4th_start'] = member['startDate']
                except:
                    pass
                try:
                    info['4th_emp'] = member['numberOfEmployees'][0]['value']
                except:
                    pass
                try:
                    info['4th_rev'] = member['financials'][0] \
                        ['yearlyRevenue'][0]['value']
                except:
                    pass
                try:
                    info['4th_parent_duns'] = member['corporateLinkage'] \
                        ['parent']['duns']
                except:
                    pass
                try:
                    items = member['corporateLinkage']['familytreeRolesPlayed']
                    num = len(items)
                    counter = 0
                    cell = ''
                    for item in items:
                        cell += items[counter]['description']
                        counter += 1
                        if counter < num:
                            cell += ', '
                    info['4th_roles'] = cell
                except:
                    pass
    except:
        pass

    return info

# This provides information on a company's parents
def parent_lookup(duns, token):
    info = {}
    serviceurl = 'https://plus.dnb.com/v1/data/duns/' + duns + '?'
    auth = 'Bearer ' + token
    headers = {'accept' : 'application/json;charset=utf-8',
        'authorization' : auth}
    params = {
    'blockIDs' : 'hierarchyconnections_L1_v1',
    'tradeUp' : 'hq',
    'customerReference' : 'katsumi_family_lookup'}
    lookup_string = serviceurl + urllib.parse.urlencode(params)
    conn = http.client.HTTPSConnection("plus.dnb.com")
    conn.request("GET", lookup_string, headers=headers)
    r = conn.getresponse()
    r = r.read().decode()
    try:
        js = json.loads(r)
    except:
        return info
    try:
        info['id'] = js['transactionDetail']['transactionID']
    except:
        info['id'] = ' '
    try:
        info['timestamp'] = js['transactionDetail']['transactionTimestamp']
    except:
        info['timestamp'] = ' '
    try:
        info['obligor_name'] = js['organization']['primaryName']
    except:
        info['obligor_name'] = ' '
    try:
        roles = js['organization']['corporateLinkage'] \
            ['familytreeRolesPlayed']
        num = len(roles)
        counter = 0
        cell = ''
        for role in roles:
            cell += js['organization']['corporateLinkage'] \
                ['familytreeRolesPlayed'][counter]['description']
            counter += 1
            if counter < num:
                cell += ', '
        info['obligor_role'] = cell
    except:
        info['obligor_role'] = ' '
    try:
        info['obligor_rank'] = js['organization']['corporateLinkage'] \
            ['hierarchyLevel']
    except:
        info['obligor_rank'] = ' '
    try:
        info['global_hq_duns'] = js['organization']['corporateLinkage'] \
            ['globalUltimate']['duns']
    except:
        info['global_hq_duns'] = ' '
    try:
        info['global_hq_name'] = js['organization']['corporateLinkage'] \
            ['globalUltimate']['primaryName']
    except:
        info['global_hq_name'] = ' '
    try:
        info['domestic_hq_duns'] = js['organization']['corporateLinkage'] \
            ['domesticUltimate']['duns']
    except:
        info['domestic_hq_duns'] = ' '
    try:
        info['domestic_hq_name'] = js['organization']['corporateLinkage'] \
            ['domesticUltimate']['primaryName']
    except:
        info['domestic_hq_name'] = ' '
    try:
        info['parent_duns'] = js['organization']['corporateLinkage'] \
            ['parent']['duns']
    except:
        info['parent_duns'] = ' '
    try:
        info['parent_name'] = js['organization']['corporateLinkage'] \
            ['parent']['primaryName']
    except:
        info['parent_name'] = ' '
    return info

# Runs KYC on input from an Excel file and updates the file
def update(file):
    key = dnb()
    token = get_token(key)
    wb = openpyxl.load_workbook(file)
    ws = wb['KYC']
    r = 4
    c = 1
    end = False
    while end == False:
        if isinstance(ws.cell(row = r, column = c).value, str):
            company = ws.cell(row = r, column = c).value
        else: break
        if isinstance(ws.cell(row = r, column = c+1).value, str) and \
            ws.cell(row = r, column = c+12).value == None:
            address = ws.cell(row = r, column = c+1).value
            location = address_parser(address)
            company = company_lookup(company, location, token)
            duns = company['1st_duns']
#            family = parent_lookup(duns, token)
            family = family_lookup(duns, token)

            ws.cell(row = r, column = c+2, value = \
                location['formatted_address'])
            ws.cell(row = r, column = c+3, value = \
                location['street_number'])
            ws.cell(row = r, column = c+4, value = \
                location['street'])
            ws.cell(row = r, column = c+5, value = \
                location['city'])
            ws.cell(row = r, column = c+6, value = \
                location['county'])
            ws.cell(row = r, column = c+7, value = \
                location['state'])
            ws.cell(row = r, column = c+8, value = \
                location['postal_code'])
            ws.cell(row = r, column = c+9, value = \
                location['country_abbrev'])
            ws.cell(row = r, column = c+10, value = \
                location['lat'])
            ws.cell(row = r, column = c+11, value = \
                location['lng'])
            try:
                ws.cell(row = r, column = c+12, value = company['id'])
            except:
                ws.cell(row = r, column = c+12, value = ' ')
            try:
                ws.cell(row = r, column = c+13, value = company['timestamp'])
            except:
                ws.cell(row = r, column = c+13, value = ' ')
            try:
                ws.cell(row = r, column = c+14, value = company['match_number'])
            except:
                ws.cell(row = r, column = c+14, value = ' ')
            try:
                ws.cell(row = r, column = c+15, value = \
                    company['1st_confidence_code'])
            except:
                ws.cell(row = r, column = c+15, value = ' ')
            try:
                ws.cell(row = r, column = c+16, value = company['1st_name'])
            except:
                ws.cell(row = r, column = c+16, value = ' ')
            try:
                ws.cell(row = r, column = c+17, value = \
                    company['1st_tradestyle_names'])
            except:
                ws.cell(row = r, column = c+17, value = ' ')
            try:
                ws.cell(row = r, column = c+18, value = company['1st_duns'])
            except:
                ws.cell(row = r, column = c+18, value = ' ')
            try:
                ws.cell(row = r, column = c+19, value = company['1st_status'])
            except:
                ws.cell(row = r, column = c+19, value = ' ')
            try:
                ws.cell(row = r, column = c+20, value = company['1st_family'])
            except:
                ws.cell(row = r, column = c+20, value = ' ')
            try:
                ws.cell(row = r, column = c+21, value = \
                    company['1st_registration_num'])
            except:
                ws.cell(row = r, column = c+21, value = ' ')
            try:
                ws.cell(row = r, column = c+22, value = \
                    company['1st_registration_type'])
            except:
                ws.cell(row = r, column = c+22, value = ' ')
            try:
                ws.cell(row = r, column = c+23, value = company['1st_website'])
            except:
                ws.cell(row = r, column = c+23, value = ' ')
            try:
                ws.cell(row = r, column = c+24, value = \
                    company['1st_telephone'])
            except:
                ws.cell(row = r, column = c+24, value = ' ')
            try:
                ws.cell(row = r, column = c+25, value = \
                    company['1st_principals'])
            except:
                ws.cell(row = r, column = c+25, value = ' ')
            try:
                ws.cell(row = r, column = c+26, value = \
                    company['1st_street'])
            except:
                ws.cell(row = r, column = c+26, value = ' ')
            try:
                ws.cell(row = r, column = c+27, value = company['1st_city'])
            except:
                ws.cell(row = r, column = c+27, value = ' ')
            try:
                ws.cell(row = r, column = c+28, value = company['1st_state'])
            except:
                ws.cell(row = r, column = c+28, value = ' ')
            try:
                ws.cell(row = r, column = c+29, value = company['1st_postal'])
            except:
                ws.cell(row = r, column = c+29, value = ' ')
            try:
                ws.cell(row = r, column = c+30, value = company['1st_country'])
            except:
                ws.cell(row = r, column = c+30, value = ' ')
            try:
                ws.cell(row = r, column = c+31, value = \
                    company['2nd_confidence_code'])
            except:
                ws.cell(row = r, column = c+31, value = ' ')
            try:
                ws.cell(row = r, column = c+32, value = company['2nd_name'])
            except:
                ws.cell(row = r, column = c+32, value = ' ')
            try:
                ws.cell(row = r, column = c+33, value = \
                    company['2nd_tradestyle_names'])
            except:
                ws.cell(row = r, column = c+33, value = ' ')
            try:
                ws.cell(row = r, column = c+34, value = company['2nd_duns'])
            except:
                ws.cell(row = r, column = c+34, value = ' ')
            try:
                ws.cell(row = r, column = c+35, value = company['2nd_status'])
            except:
                ws.cell(row = r, column = c+35, value = ' ')
            try:
                ws.cell(row = r, column = c+36, value = company['2nd_family'])
            except:
                ws.cell(row = r, column = c+36, value = ' ')
            try:
                ws.cell(row = r, column = c+37, value = \
                    company['2nd_registration_num'])
            except:
                ws.cell(row = r, column = c+37, value = ' ')
            try:
                ws.cell(row = r, column = c+38, value = \
                    company['2nd_registration_type'])
            except:
                ws.cell(row = r, column = c+38, value = ' ')
            try:
                ws.cell(row = r, column = c+39, value = company['2nd_website'])
            except:
                ws.cell(row = r, column = c+39, value = ' ')
            try:
                ws.cell(row = r, column = c+40, value = \
                   company['2nd_telephone'])
            except:
                ws.cell(row = r, column = c+40, value = ' ')
            try:
                ws.cell(row = r, column = c+41, value = \
                    company['2nd_principals'])
            except:
                ws.cell(row = r, column = c+41, value = ' ')
            try:
                ws.cell(row = r, column = c+42, value = \
                    company['2nd_street'])
            except:
                ws.cell(row = r, column = c+42, value = ' ')
            try:
                ws.cell(row = r, column = c+43, value = company['2nd_city'])
            except:
                ws.cell(row = r, column = c+43, value = ' ')
            try:
                ws.cell(row = r, column = c+44, value = company['2nd_state'])
            except:
                ws.cell(row = r, column = c+44, value = ' ')
            try:
                ws.cell(row = r, column = c+45, value = company['2nd_postal'])
            except:
                ws.cell(row = r, column = c+45, value = ' ')
            try:
                ws.cell(row = r, column = c+46, value = company['2nd_country'])
            except:
                ws.cell(row = r, column = c+46, value = ' ')
            try:
                ws.cell(row = r, column = c+47, value = family['id'])
            except:
                ws.cell(row = r, column = c+47, value = ' ')
            try:
                ws.cell(row = r, column = c+48, value = family['timestamp'])
            except:
                ws.cell(row = r, column = c+48, value = ' ')
            try:
                ws.cell(row = r, column = c+49, value = family['1st_name'])
            except:
                ws.cell(row = r, column = c+49, value = ' ')
            try:
                ws.cell(row = r, column = c+50, value = family['1st_trade'])
            except:
                ws.cell(row = r, column = c+50, value = ' ')
            try:
                ws.cell(row = r, column = c+51, value = family['1st_duns'])
            except:
                ws.cell(row = r, column = c+51, value = ' ')
            try:
                ws.cell(row = r, column = c+52, value = family['1st_sic'])
            except:
                ws.cell(row = r, column = c+52, value = ' ')
            try:
                ws.cell(row = r, column = c+53, value = family['1st_level'])
            except:
                ws.cell(row = r, column = c+53, value = ' ')
            try:
                ws.cell(row = r, column = c+54, value = family['1st_start'])
            except:
                ws.cell(row = r, column = c+54, value = ' ')
            try:
                ws.cell(row = r, column = c+55, value = family['1st_rev'])
            except:
                ws.cell(row = r, column = c+55, value = ' ')
            try:
                ws.cell(row = r, column = c+56, value = family['1st_emp'])
            except:
                ws.cell(row = r, column = c+56, value = ' ')


            try:
                ws.cell(row = r, column = c+57, value = family['2nd_name'])
            except:
                ws.cell(row = r, column = c+57, value = ' ')
            try:
                ws.cell(row = r, column = c+58, value = family['2nd_trade'])
            except:
                ws.cell(row = r, column = c+58, value = ' ')
            try:
                ws.cell(row = r, column = c+59, value = family['2nd_duns'])
            except:
                ws.cell(row = r, column = c+59, value = ' ')
            try:
                ws.cell(row = r, column = c+60, value = family['2nd_sic'])
            except:
                ws.cell(row = r, column = c+60, value = ' ')
            try:
                ws.cell(row = r, column = c+61, value = family['2nd_level'])
            except:
                ws.cell(row = r, column = c+61, value = ' ')
            try:
                ws.cell(row = r, column = c+62, value = family['2nd_start'])
            except:
                ws.cell(row = r, column = c+62, value = ' ')
            try:
                ws.cell(row = r, column = c+63, value = family['2nd_rev'])
            except:
                ws.cell(row = r, column = c+63, value = ' ')
            try:
                ws.cell(row = r, column = c+64, value = family['2nd_emp'])
            except:
                ws.cell(row = r, column = c+64, value = ' ')

            try:
                ws.cell(row = r, column = c+65, value = family['3rd_name'])
            except:
                ws.cell(row = r, column = c+65, value = ' ')
            try:
                ws.cell(row = r, column = c+66, value = family['3rd_trade'])
            except:
                ws.cell(row = r, column = c+66, value = ' ')
            try:
                ws.cell(row = r, column = c+67, value = family['3rd_duns'])
            except:
                ws.cell(row = r, column = c+67, value = ' ')
            try:
                ws.cell(row = r, column = c+68, value = family['3rd_sic'])
            except:
                ws.cell(row = r, column = c+68, value = ' ')
            try:
                ws.cell(row = r, column = c+69, value = family['3rd_level'])
            except:
                ws.cell(row = r, column = c+69, value = ' ')
            try:
                ws.cell(row = r, column = c+70, value = family['3rd_start'])
            except:
                ws.cell(row = r, column = c+70, value = ' ')
            try:
                ws.cell(row = r, column = c+71, value = family['3rd_rev'])
            except:
                ws.cell(row = r, column = c+71, value = ' ')
            try:
                ws.cell(row = r, column = c+72, value = family['3rd_emp'])
            except:
                ws.cell(row = r, column = c+72, value = ' ')

            try:
                ws.cell(row = r, column = c+73, value = family['4th_name'])
            except:
                ws.cell(row = r, column = c+73, value = ' ')
            try:
                ws.cell(row = r, column = c+74, value = family['4th_trade'])
            except:
                ws.cell(row = r, column = c+74, value = ' ')
            try:
                ws.cell(row = r, column = c+75, value = family['4th_duns'])
            except:
                ws.cell(row = r, column = c+75, value = ' ')
            try:
                ws.cell(row = r, column = c+76, value = family['4th_sic'])
            except:
                ws.cell(row = r, column = c+76, value = ' ')
            try:
                ws.cell(row = r, column = c+77, value = family['4th_level'])
            except:
                ws.cell(row = r, column = c+77, value = ' ')
            try:
                ws.cell(row = r, column = c+78, value = family['4th_start'])
            except:
                ws.cell(row = r, column = c+78, value = ' ')
            try:
                ws.cell(row = r, column = c+79, value = family['4th_rev'])
            except:
                ws.cell(row = r, column = c+79, value = ' ')
            try:
                ws.cell(row = r, column = c+80, value = family['4th_emp'])
            except:
                ws.cell(row = r, column = c+80, value = ' ')

        r += 1
        if r > 200: end = True
    wb.save(file)


# This one works - keep as backup
def test8():
    headers = {
        'Authorization': 'Basic ZmMyNGY0YzU3MzA4NDQ5Nzk3M2Q2OWRkMjJhNTllMmJkNmY2ZTZmZTRhYjc0YmJiOTY1Zjc1OTUyYTgzYThmZTplMDE2NWU2MWM5ZDQ0ZWU0YTAxMTA0MzRkNWJiOWU2MDE1YzVhMzkzZGQ1ODRlMTg5YjlmNTY5MTQ0YjVkYTcw',
        'Content-Type': 'application/json',
    }

    data = '{ "grant_type" : "client_credentials" }'

    response = requests.post('https://plus.dnb.com/v2/token', headers=headers, data=data)
    print(response)



if __name__ == '__main__':
    update('KYC.xlsx')
